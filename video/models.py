import operator

import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.utils import timezone
from openpyxl import load_workbook
from toolkit.models import CCEAuditModel
from organization.models import Party, Office, Organization
from people.models import PersonRole
from taxonomy.models import Tag, ElectionYear, VideoFormat
from utils.locations import COUNTRIES
from video.reports import VideoReports


class Video(CCEAuditModel):
    IMAGE = 'image'
    POSITIVE = 'positive'
    NEGATIVE = "negative"
    POSITIVE_IMAGE = 'positive_image'
    POSITIVE_ISSUE = 'positive_issue'
    NEGATIVE_IMAGE = 'negative_image'
    NEGATIVE_ISSUE = 'negative_issue'
    ISSUE = 'issue'
    OTHER = 'other'

    COMMUNICATION_TYPE_CHOICES = (
        (IMAGE, 'Image'),
        (POSITIVE, 'Positive'),
        (NEGATIVE, 'Negative'),
        (POSITIVE_IMAGE, 'Positive Image'),
        (POSITIVE_ISSUE, 'Positive Issue'),
        (NEGATIVE_IMAGE, 'Negative Image'),
        (NEGATIVE_ISSUE, 'Negative Issue'),
        (ISSUE, 'Issue'),
        (OTHER, 'other'),
    )

    PROGRAM_TYPE_CHOICES = (
        ('commercial', 'Commercial'),
        ('convention', 'Convention'),
        ('debate', 'Debate'),
        ('interview', 'Interview'),
        ('news', 'News'),
        ('program', 'Program'),
        ('public_service_announcement', 'Public Service Announcement'),
        ('rally', 'Rally'),
        ('speech', 'Speech'),
        ('other', 'Other'),
    )

    STATE_CHOICES = (
        ('alabama', 'Alabama'),
        ('alaska', 'Alaska'),
        ('american_samoa', 'American Samoa'),
        ('arizona', 'Arizona'),
        ('arkansas', 'Arkansas'),
        ('california', 'California'),
        ('colorado', 'Colorado'),
        ('connecticut', 'Connecticut'),
        ('delaware', 'Delaware'),
        ('washington_district_of_columbia', 'Washington District of Columbia'),
        ('federated_states_of_micronesia', 'Federated States of Micronesia'),
        ('florida', 'Florida'),
        ('georgia', 'Georgia'),
        ('guam', 'Guam'),
        ('hawaii', 'Hawaii'),
        ('idaho', 'Idaho'),
        ('illinois', 'Illinois'),
        ('indiana', 'Indiana'),
        ('iowa', 'Iowa'),
        ('kansas', 'Kansas'),
        ('kentucky', 'Kentucky'),
        ('louisiana', 'Louisiana'),
        ('maine', 'Maine'),
        ('marshall_islands', 'Marshall Islands'),
        ('maryland', 'Maryland'),
        ('massachusetts', 'Massachusetts'),
        ('michigan', 'Michigan'),
        ('minnesota', 'Minnesota'),
        ('mississippi', 'Mississippi'),
        ('missouri', 'Missouri'),
        ('montana', 'Montana'),
        ('nebraska', 'Nebraska'),
        ('nevada', 'Nevada'),
        ('new_hampshire', 'New Hampshire'),
        ('new_jersey', 'New Jersey'),
        ('new_mexico', 'New Mexico'),
        ('new_york', 'New York'),
        ('north_carolina', 'North Carolina'),
        ('north_dakota', 'North Dakota'),
        ('northern_mariana_islands', 'Northern Mariana Islands'),
        ('ohio', 'Ohio'),
        ('oklahoma', 'Oklahoma'),
        ('ontario', 'Ontario'),
        ('oregon', 'Oregon'),
        ('palau', 'Palau'),
        ('pennsylvania', 'Pennsylvania'),
        ('puerto_rico', 'Puerto Rico'),
        ('rhode_island', 'Rhode Island'),
        ('rhode_island', 'Rhode Island'),
        ('sint_marten', 'Sint Maarten'),
        ('south_carikuba', 'South Carolina'),
        ('south_dakota', 'South Dakota'),
        ('tennessee', 'Tennessee'),
        ('texas', 'Texas'),
        ('utah', 'Utah'),
        ('vermont', 'Vermont'),
        ('virgin_islands', 'Virgin Islands'),
        ('virginia', 'Virginia'),
        ('washington', 'Washington'),
        ('west_virginia', 'West Virginia'),
        ('wisconsin', 'Wisconsin'),
        ('wyoming', 'Wyoming'),
        ('west_germany', 'West Germany'),
        ('other', 'Other'),
    )

    GENDER_CHOICES = (
        ('male', 'Male'),
        ('female', 'Female'),
        ('transgender', 'Transgender'),
        ('unknown', 'Unknown'),
    )

    COUNTRY_CHOICES = sorted(COUNTRIES, key=operator.itemgetter(0))

    database_id = models.IntegerField(unique=True)
    original_id = models.IntegerField(unique=True)
    preservation_copy = models.CharField(max_length=25, blank=True)
    political_commercial_archive = models.CharField(max_length=25, blank=True)
    slate = models.CharField(max_length=25, blank=True)
    creation_date = models.DateField(null=True)
    communication_type = models.CharField(max_length=50, choices= COMMUNICATION_TYPE_CHOICES)
    program_type = models.CharField(max_length=50, choices=PROGRAM_TYPE_CHOICES)
    election_year = models.ForeignKey(ElectionYear, related_name='ElectionYear', null=True)
    format = models.ForeignKey(VideoFormat, related_name='VideoFormat', null=True)
    agency = models.CharField(max_length=50, blank=True, null=True)
    length = models.CharField(max_length=25, blank=True, null=True)
    begin_time = models.CharField(max_length=25, blank=True, null=True)
    first_name = models.CharField(max_length=50, blank=True, null=True)
    last_name = models.CharField(max_length=50, blank=True, null=True)
    organization = models.CharField(max_length=50, blank=True, null=True)
    role = models.ForeignKey(PersonRole, related_name='PersonRole', null=True)
    country = models.CharField(max_length=50, choices=COUNTRY_CHOICES)
    party = models.ForeignKey(Party, related_name='Party', null=True)
    state = models.CharField(max_length=50, choices= STATE_CHOICES)
    office = models.ForeignKey(Office, related_name='Office', null=True)
    gender = models.CharField(max_length=10, choices= GENDER_CHOICES)
    title = models.CharField(max_length=50, blank=True, null=True)
    notes = models.CharField(max_length=1000, blank=True, null=True)
    summary = models.TextField(blank=True, null=True)
    transcript = models.TextField(blank=True, null=True)
    subject1 = models.CharField(max_length=50, blank=True, null=True)
    subject2 = models.CharField(max_length=50, blank=True, null=True)
    subject3 = models.CharField(max_length=50, blank=True, null=True)
    cataloged_date = models.DateTimeField(auto_now_add=False, null=True)
    donor = models.CharField(max_length=50, blank=True, null=True)
    licence = models.CharField(max_length=50, blank=True, null=True)
    cataloger = models.CharField(max_length=50, blank=True, null=True)
    tags = models.ManyToManyField(Tag)

    reports = VideoReports()

    def __unicode__(self):
        return "%s, %s" % (self.database_id,self.preservation_copy)

    @classmethod
    def import_videos(cls, xlsxfile):
        line_number = 2
        created_list = []
        updated_list = []
        error_list = []

        if xlsxfile.content_type != u'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            error_list.append('Please submit a valid xlsx file.')
            return created_list, updated_list, error_list

        workbook = load_workbook(xlsxfile, read_only=True)
        sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])

        def get_string_value(value):
            return unicode(value).strip() if value else ''

        def get_int_value(value):
            return int(value) if value else ''

        def get_date_value(value):
            if value is None:
                return None
            if value == u'0000-00-00':
                return u"0000-00-00"
            if isinstance(value, datetime.datetime):
                return value
            if '-' in value and value != u'0000-00-00':
                return datetime.datetime.strptime(value, "%Y-%m-%d")
            else:
                raise Exception(
                    'Value "%s" has an invalid date format. Please enter the date in the format YYYY-MM-DD' % value
                )

        def get_election_year(value):
            if value is '' or value is None:
                return None
            election_year = ElectionYear.objects.filter(year=value)
            if not election_year.exists():
                raise ObjectDoesNotExist('Election Year "%s" does not exist' % value)
            else:
                return ElectionYear.objects.get(year=value)

        def get_format(value):
            if value is '' or value is None:
                return None
            format_name = VideoFormat.objects.filter(name=value)
            if not format_name.exists():
                raise ObjectDoesNotExist('Video Format Name "%s" does not exist' % value)
            else:
                return VideoFormat.objects.get(name=value)

        def get_role(value):
            if value is '' or value is None:
                return None
            person_role = PersonRole.objects.filter(name=value)
            if not person_role.exists():
                raise ObjectDoesNotExist('person role name "%s" does not exist' % value)
            else:
                return PersonRole.objects.get(name=value)

        def get_party(value):
            if value is '' or value is None:
                return None
            party = Party.objects.filter(name=value)
            if not party.exists():
                raise ObjectDoesNotExist('party name "%s" does not exist' % value)
            else:
                return Party.objects.get(name=value)

        def get_office(value):
            if value is '' or value is None:
                return None
            office = Office.objects.filter(name=value)
            if not office.exists():
                raise ObjectDoesNotExist(' office name "%s" does not exist' % value )
            else:
                return Office.objects.get(name=value)

        def get_datetime_value(value):
            if value is '' or value is None:
                return None
            if value == u'0000-00-00 00:00:00':
                return value
            if isinstance(value, datetime.datetime):
                return value
            if '-' in value and value != u'0000-00-00 00:00:00':
                return datetime.datetime.strftime(value, "%Y-%m-%d %H:%M:%S")
            else:
                raise Exception(
                    'Value "%s" has an invalid datetime format. Please enter the date in the format YYYY-MM-DD H-M-S' % value
                )

        def get_tags_values(value):
            if value is None or value is '':
                return None
            if ';' in value:
                result = value.split('; ')
            else:
                result = value.split(', ')
            for tag in result:
                if str(tag) is not "" and tag is not None:
                    tag_exists = Tag.objects.filter(name=tag).exists()
                    if not tag_exists:
                        raise ObjectDoesNotExist('"%s" is not a valid value for Tags' % tag)

            return result

        for video_data in sheet.iter_rows(row_offset=1):
            try:
                database_id = get_int_value(video_data[0].value)
                original_id = get_int_value(video_data[1].value)
                preservation_copy = get_string_value(video_data[2].value)
                political_commercial_archive = get_string_value(video_data[3].value)
                slate = get_string_value(video_data[4].value)
                creation_date = get_date_value(video_data[5].value)
                communication_type = get_string_value(video_data[6].value)
                program_type = get_string_value(video_data[7].value)
                election_year = get_election_year(video_data[8].value)
                format = get_format(video_data[9].value)
                agency = get_string_value(video_data[10].value)
                length = get_string_value(video_data[11].value)
                begin_time = get_string_value(video_data[12].value)
                first_name = get_string_value(video_data[13].value)
                last_name = get_string_value(video_data[14].value)
                organization = get_string_value(video_data[15].value)
                role = get_role(video_data[16].value)
                country = get_string_value(video_data[17].value)
                party = get_party(video_data[18].value)
                state = get_string_value(video_data[19].value)
                office = get_office(video_data[20].value)
                gender = get_string_value(video_data[21].value)
                title = get_string_value(video_data[22].value)
                notes = get_string_value(video_data[23].value)
                summary = get_string_value(video_data[24].value)
                transcript = get_string_value(video_data[25].value)
                subject1 = get_string_value(video_data[26].value)
                subject2 = get_string_value(video_data[27].value)
                subject3 = get_string_value(video_data[28].value)
                cataloged_date = get_datetime_value(video_data[29].value)
                donor = get_string_value(video_data[30].value)
                licence = get_string_value(video_data[31].value)
                cataloger = get_string_value(video_data[32].value)
                tags = get_tags_values(video_data[33].value)

                if Video.objects.filter(database_id=database_id).exists():
                    video = Video.objects.get(database_id=database_id)
                    video.database_id = database_id
                    video.original_id=original_id
                    video.preservation_copy=preservation_copy
                    video.political_commercial_archive=political_commercial_archive
                    video.slate=slate
                    video.creation_date=creation_date
                    video.communication_type=communication_type
                    video.program_type=program_type
                    video.election_year=election_year
                    video.format=format
                    video.agency=agency
                    video.length=length
                    video.begin_time=begin_time
                    video.first_name=first_name
                    video.last_name=last_name
                    video.organization=organization
                    video.role=role
                    video.country=country
                    video.party=party
                    video.state=state
                    video.office=office
                    video.gender=gender
                    video.title=title
                    video.notes=notes
                    video.summary=summary
                    video.transcript=transcript
                    video.subject1=subject1
                    video.subject2=subject2
                    video.subject3=subject3
                    video.cataloged_date=cataloged_date
                    video.donor=donor
                    video.licence=licence
                    video.cataloger=cataloger
                    video.save()
                    if tags is not None:
                        for tag in tags:
                            video.tags.add(Tag.objects.get(name=tag))
                    updated_list.append('Video with database id %s updated successfully' % video.database_id)
                else:
                    video = Video.objects.create(
                        database_id=database_id,
                        original_id=original_id,
                        preservation_copy=preservation_copy,
                        political_commercial_archive=political_commercial_archive,
                        slate=slate,
                        creation_date=creation_date,
                        communication_type=communication_type,
                        program_type=program_type,
                        election_year=election_year,
                        format=format,
                        agency=agency,
                        length=length,
                        begin_time=begin_time,
                        first_name=first_name,
                        last_name=last_name,
                        organization=organization,
                        role=role,
                        country=country,
                        party=party,
                        state=state,
                        office=office,
                        gender=gender,
                        title=title,
                        notes=notes,
                        summary=summary,
                        transcript=transcript,
                        subject1=subject1,
                        subject2=subject2,
                        subject3=subject3,
                        cataloged_date=cataloged_date,
                        donor=donor,
                        licence=licence,
                        cataloger=cataloger,
                    )
                    if tags is not None:
                        for tag in tags:
                            video.tags.add(Tag.objects.get(name=tag))
                    created_list.append('Video with database id %s created successfully' % video.database_id)

            except Exception as e:
                error_list.append('Error (Row: %s): %s' % (line_number, e))
            line_number += 1

        return created_list, updated_list, error_list

    def can_update(self, user_obj):
        return True

    def can_delete(self, user_obj):
        return True

    def can_create(self, user_obj):
        return True

    def can_view_list(self, user_obj):
        return True

    def can_view(self, user_obj):
        return True


class Comment(CCEAuditModel):
    video = models.ForeignKey(Video)
    name = models.CharField(max_length=25)
    comment = models.TextField(max_length=250)
    created_date = models.DateTimeField(default=timezone.now)

    def __str__(self):
        return self.comment

    def can_update(self, user_obj):
        return True

    def can_delete(self, user_obj):
        return user_obj.is_staff or self.created_by == user_obj

    def can_create(self, user_obj):
        return True

    def can_view_list(self, user_obj):
        return True

    def can_view(self, user_obj):
        return True

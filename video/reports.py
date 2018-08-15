
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


class VideoReports(object):
    VIDEO_COLUMNS =[
        (U"Database Id", 30),
        (u"Original Id", 30),
        (u"Preservation Copy", 30),
        (u"Political Commercial Archive", 30),
        (u"Slate", 30),
        (u"Creation Date", 30),
        (u"Communication Type", 30),
        (u"Program Type", 30),
        (u"Election Year", 30),
        (u"Format", 30),
        (u"Agency", 30),
        (u"Length", 30),
        (u"Begin Time", 30),
        (u"First Name", 30),
        (u"Last Name", 30),
        (u"Organization", 30),
        (u"Role", 30),
        (u"Country", 30),
        (u"Party", 30),
        (u"State", 30),
        (u"Office", 30),
        (u"Gender", 30),
        (u"Title", 30),
        (u"Notes", 30),
        (u"Summary", 30),
        (u"Transcript", 30),
        (u"Subject1", 30),
        (u"Subject2", 30),
        (u"Subject3", 30),
        (u"Cataloged Date", 30),
        (u"Donor", 30),
        (u"Licence", 30),
        (u"Cataloger", 30),
        (u"Tags", 30),
    ]

    def video_xlsx_export(cls, qs, form):
        from video.models import Video
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="videos_export.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Video Cataloger"

        row_num = 0
        columns = cls.VIDEO_COLUMNS

        for col_num in xrange(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]

            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for video in Video.objects.filter(pk__in=qs.values_list('pk', flat=True)):
            row_num += 1
            row = [
                video.database_id,
                video.original_id,
                video.preservation_copy,
                video.political_commercial_archive,
                video.slate,
                video.creation_date,
                video.communication_type,
                video.program_type,
                video.election_year.year if video.election_year else "",
                video.format.name if video.format else "",
                video.agency,
                video.length,
                video.begin_time,
                video.first_name,
                video.last_name,
                video.organization,
                video.role.name if video.format else "",
                video.country,
                video.party.name if video.party else "",
                video.state,
                video.office.name if video.office else "",
                video.gender,
                video.title,
                video.notes,
                video.summary,
                video.transcript,
                video.subject1,
                video.subject2,
                video.subject3,
                video.cataloged_date,
                video.donor,
                video.licence,
                video.cataloger,
                ", ".join(t.name for t in video.tags.all()),
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

        wb.save(response)
        return response
